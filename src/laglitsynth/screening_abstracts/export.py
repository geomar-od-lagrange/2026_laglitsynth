"""CSV and XLSX export of stage 3 verdicts for human review.

Single subcommand ``screening-abstracts-export`` with ``--format csv|xlsx``.

CSV mode writes a flat ``review.csv`` a reviewer opens in Excel, Numbers, or
Keynote. Two empty columns (``reviewer_decision``, ``reviewer_reason``) are
filled in the spreadsheet; a stage-4 ingestor will read them back.

XLSX mode writes a workbook with an ``Index`` sheet plus one tab per included
work. Per-work tabs use a vertical ``Field | Value`` layout so abstract and
``raw_response`` wrap vertically without horizontal scrolling. The XLSX
output also pulls the screening criterion (the user prompt) from
``screening-meta.json`` so the reviewer sees the same question the LLM saw,
verbatim. The LLM verdict block is collapsed by default so the reviewer
forms an opinion before peeking.

``--n-subset`` + ``--subset-seed`` draw a reproducible random sample;
``n_subset >= len(verdicts)`` (or unset) emits the full set in verdict order.
``--n-subset`` and ``--subset-seed`` are valid only with ``--format xlsx``.
"""

from __future__ import annotations

import argparse
import csv
import json
import random
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.worksheet import Worksheet

from laglitsynth.catalogue_fetch.models import Work
from laglitsynth.io import read_jsonl
from laglitsynth.screening_abstracts.models import ScreeningMeta, ScreeningVerdict

# ── CSV ───────────────────────────────────────────────────────────────────────

COLUMNS: tuple[str, ...] = (
    "work_id",
    "title",
    "doi",
    "publication_year",
    "abstract",
    "relevance_score",
    "llm_reason",
    "reviewer_decision",
    "reviewer_reason",
    "raw_response",
)


def build_row(verdict: ScreeningVerdict, work: Work) -> dict[str, str]:
    """One CSV row. Missing values render as empty cells."""
    return {
        "work_id": verdict.work_id,
        "title": work.title or "",
        "doi": work.doi or "",
        "publication_year": (
            str(work.publication_year) if work.publication_year is not None else ""
        ),
        "abstract": work.abstract or "",
        "relevance_score": (
            str(verdict.relevance_score)
            if verdict.relevance_score is not None
            else ""
        ),
        "llm_reason": verdict.reason or "",
        "reviewer_decision": "",
        "reviewer_reason": "",
        "raw_response": verdict.raw_response or "",
    }


def export_review_csv(
    verdicts_path: Path,
    catalogue_path: Path,
    output_path: Path,
) -> int:
    """Write the review CSV and return the number of data rows.

    Raises ``ValueError`` naming the first ``work_id`` present in the
    verdicts file but absent from the catalogue.
    """
    catalogue: dict[str, Work] = {w.id: w for w in read_jsonl(catalogue_path, Work)}
    rows: list[dict[str, str]] = []
    for verdict in read_jsonl(verdicts_path, ScreeningVerdict):
        work = catalogue.get(verdict.work_id)
        if work is None:
            raise ValueError(
                f"work_id {verdict.work_id!r} in {verdicts_path} not found in "
                f"catalogue {catalogue_path}"
            )
        rows.append(build_row(verdict, work))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(COLUMNS))
        writer.writeheader()
        writer.writerows(rows)
    return len(rows)


# ── XLSX ──────────────────────────────────────────────────────────────────────

_SHEET_NAME_MAX_LEN = 31
_FORBIDDEN_SHEET_CHARS = set("/\\?*[]:")
_DEFAULT_SUBSET_SEED = 0

_REVIEWER_SCORE_PLACEHOLDER = (
    "<insert relevance between 0% and 100% here — score based on info "
    "provided above only>"
)
_REVIEWER_REASON_PLACEHOLDER = (
    "<give short (max 3 sentences) reason for the relevance score>"
)
_SCORING_INSTRUCTIONS = "Score 0% (not relevant) to 100% (perfectly relevant)"

_HYPERLINK_FONT = Font(color="0563C1", underline="single")
_BOLD = Font(bold=True)
_HEADER_FILL = PatternFill("solid", fgColor="E0E0E0")


def sample_verdicts(
    verdicts: list[ScreeningVerdict],
    n_subset: int | None,
    seed: int,
) -> list[ScreeningVerdict]:
    """Return a subset of verdicts, preserving their original order.

    ``n_subset is None`` or ``n_subset >= len(verdicts)`` returns the
    full list. Otherwise draw ``n_subset`` work_ids uniformly with
    ``random.Random(seed)`` and emit the matching verdicts in the
    order they appear in ``verdicts``.
    """
    if n_subset is None or n_subset >= len(verdicts):
        return list(verdicts)
    rng = random.Random(seed)
    chosen_ids = set(rng.sample([v.work_id for v in verdicts], k=n_subset))
    return [v for v in verdicts if v.work_id in chosen_ids]


def short_work_id(work_id: str) -> str:
    """Trailing OpenAlex id, sanitised for Excel sheet naming.

    Strips host/path prefix, replaces characters Excel forbids in
    sheet names with ``_``, truncates to 31 chars.
    """
    name = work_id.rsplit("/", 1)[-1] if "/" in work_id else work_id
    name = "".join("_" if c in _FORBIDDEN_SHEET_CHARS else c for c in name)
    name = name.strip("'")
    return name[:_SHEET_NAME_MAX_LEN] or "work"


def _unique_sheet_name(base: str, used: set[str]) -> str:
    if base not in used:
        used.add(base)
        return base
    for i in range(2, 10**6):
        suffix = f"_{i}"
        candidate = f"{base[: _SHEET_NAME_MAX_LEN - len(suffix)]}{suffix}"
        if candidate not in used:
            used.add(candidate)
            return candidate
    raise RuntimeError(f"Could not resolve unique sheet name for {base!r}")


def _authors_string(work: Work) -> str:
    names = [a.author.display_name for a in work.authorships]
    return ", ".join(names)


def _journal_name(work: Work) -> str:
    if work.primary_location and work.primary_location.source:
        return work.primary_location.source.display_name or ""
    return ""


def _openalex_external_url(work_id: str) -> str:
    """openalex IDs come back as full URLs already; just normalise."""
    if work_id.startswith("http://") or work_id.startswith("https://"):
        return work_id
    trailing = work_id.rsplit("/", 1)[-1]
    return f"https://openalex.org/{trailing}"


def _doi_url(doi: str) -> str:
    if doi.startswith("http://") or doi.startswith("https://"):
        return doi
    return f"https://doi.org/{doi}"


def _percent_value(score: int | None) -> float | None:
    """Convert a 0-100 integer LLM score into the 0.0-1.0 range Excel
    multiplies by 100 for ``0%`` number-format display."""
    if score is None:
        return None
    return score / 100.0


def build_index_sheet(
    ws: Worksheet, items: list[tuple[ScreeningVerdict, Work, str]]
) -> None:
    """Index sheet with reviewer-identity header rows above the table.

    Layout::

        A1: reviewer_name      B1: <enter reviewer name here>
        A2: reviewer_email     B2: <enter reviewer email here>
        A3: review_date        B3: <YYYY-MM-DD>
        A4: (blank)
        A5: header row (frozen)
        A6..: data rows
    """
    ws.title = "Index"

    ws.cell(row=1, column=1, value="reviewer_name").font = _BOLD
    ws.cell(row=1, column=2, value="<enter reviewer name here>")
    ws.cell(row=2, column=1, value="reviewer_email").font = _BOLD
    ws.cell(row=2, column=2, value="<enter reviewer email here>")
    ws.cell(row=3, column=1, value="review_date").font = _BOLD
    ws.cell(row=3, column=2, value="<YYYY-MM-DD>")

    headers = (
        "work_id",
        "title",
        "authors",
        "journal",
        "year",
        "llm_score",
        "llm_reason",
        "sheet",
    )
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = _BOLD
        cell.fill = _HEADER_FILL

    for row_offset, (verdict, work, sheet_name) in enumerate(items):
        row = 6 + row_offset
        ws.cell(row=row, column=1, value=verdict.work_id)
        ws.cell(row=row, column=2, value=work.title or "")
        ws.cell(row=row, column=3, value=_authors_string(work))
        ws.cell(row=row, column=4, value=_journal_name(work))
        ws.cell(row=row, column=5, value=work.publication_year)
        score_cell = ws.cell(row=row, column=6, value=_percent_value(verdict.relevance_score))
        score_cell.number_format = "0%"
        ws.cell(row=row, column=7, value=verdict.reason or "")
        link_cell = ws.cell(row=row, column=8, value=sheet_name)
        link_cell.hyperlink = Hyperlink(
            ref=link_cell.coordinate,
            location=f"'{sheet_name}'!A1",
            display=sheet_name,
        )
        link_cell.font = _HYPERLINK_FONT

    # Freeze identity rows + table header.
    ws.freeze_panes = "A6"
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 60
    ws.column_dimensions["H"].width = 20


def build_work_sheet(
    ws: Worksheet,
    verdict: ScreeningVerdict,
    work: Work,
    *,
    criterion: str,
    llm_meta: dict[str, object],
) -> None:
    """Per-work sheet: bibliographic block, criterion + scoring, LLM details (collapsed).

    ``criterion`` is the screening prompt rendered verbatim so the
    reviewer scores against the same question the LLM saw.
    ``llm_meta`` carries the model/temperature/prompt_sha256 from
    ScreeningMeta.llm so the reviewer can audit the LLM's run.
    """
    wrap_top_left = Alignment(wrap_text=True, vertical="top", horizontal="left")
    top_aligned = Alignment(vertical="top")

    # Row 1: back-to-Index hyperlink
    back_cell = ws.cell(row=1, column=1, value="← back to Index")
    back_cell.hyperlink = Hyperlink(
        ref=back_cell.coordinate,
        location="'Index'!A1",
        display="← back to Index",
    )
    back_cell.font = _HYPERLINK_FONT

    # Row 2: blank

    # Bibliographic block (rows 3-9)
    biblio_rows: list[tuple[str, object]] = [
        ("title", work.title or ""),
        ("authors", _authors_string(work)),
        ("journal", _journal_name(work)),
        ("publication_year", work.publication_year),
        ("doi", None),  # filled separately to attach a hyperlink
        ("openalex", None),  # ditto
        ("abstract", work.abstract or ""),
    ]
    for offset, (field, value) in enumerate(biblio_rows):
        row = 3 + offset
        f_cell = ws.cell(row=row, column=1, value=field)
        f_cell.font = _BOLD
        f_cell.alignment = top_aligned
        if field == "doi":
            doi = work.doi or ""
            v_cell = ws.cell(row=row, column=2, value=doi)
            if doi:
                v_cell.hyperlink = Hyperlink(
                    ref=v_cell.coordinate,
                    target=_doi_url(doi),
                    display=doi,
                )
                v_cell.font = _HYPERLINK_FONT
            v_cell.alignment = wrap_top_left
        elif field == "openalex":
            url = _openalex_external_url(work.id)
            v_cell = ws.cell(row=row, column=2, value=url)
            v_cell.hyperlink = Hyperlink(
                ref=v_cell.coordinate,
                target=url,
                display=url,
            )
            v_cell.font = _HYPERLINK_FONT
            v_cell.alignment = wrap_top_left
        else:
            v_cell = ws.cell(row=row, column=2, value=value)
            v_cell.alignment = wrap_top_left

    # Row 10: blank

    # Criterion and reviewer scoring (rows 11-14)
    crit_label = ws.cell(row=11, column=1, value="criterion")
    crit_label.font = _BOLD
    crit_label.alignment = top_aligned
    crit_value = ws.cell(row=11, column=2, value=criterion)
    crit_value.alignment = wrap_top_left

    instr_label = ws.cell(row=12, column=1, value="scoring_instructions")
    instr_label.font = _BOLD
    instr_label.alignment = top_aligned
    instr_value = ws.cell(row=12, column=2, value=_SCORING_INSTRUCTIONS)
    instr_value.alignment = wrap_top_left

    rs_label = ws.cell(row=13, column=1, value="reviewer_score")
    rs_label.font = _BOLD
    rs_label.alignment = top_aligned
    rs_value = ws.cell(row=13, column=2, value=_REVIEWER_SCORE_PLACEHOLDER)
    rs_value.alignment = wrap_top_left

    rr_label = ws.cell(row=14, column=1, value="reviewer_reason")
    rr_label.font = _BOLD
    rr_label.alignment = top_aligned
    rr_value = ws.cell(row=14, column=2, value=_REVIEWER_REASON_PLACEHOLDER)
    rr_value.alignment = wrap_top_left

    # Row 15: blank

    # LLM details — header on row 16 visible, rows 17-22 collapsed.
    llm_header = ws.cell(row=16, column=1, value="LLM details (expand to peek)")
    llm_header.font = _BOLD
    llm_header.alignment = top_aligned

    llm_score_label = ws.cell(row=17, column=1, value="llm_score")
    llm_score_label.font = _BOLD
    llm_score_label.alignment = top_aligned
    llm_score_value = ws.cell(
        row=17, column=2, value=_percent_value(verdict.relevance_score)
    )
    llm_score_value.number_format = "0%"
    llm_score_value.alignment = wrap_top_left

    llm_reason_label = ws.cell(row=18, column=1, value="llm_reason")
    llm_reason_label.font = _BOLD
    llm_reason_label.alignment = top_aligned
    llm_reason_value = ws.cell(row=18, column=2, value=verdict.reason or "")
    llm_reason_value.alignment = wrap_top_left

    llm_model_label = ws.cell(row=19, column=1, value="llm_model")
    llm_model_label.font = _BOLD
    llm_model_label.alignment = top_aligned
    ws.cell(row=19, column=2, value=str(llm_meta.get("model", "")))

    llm_temp_label = ws.cell(row=20, column=1, value="llm_temperature")
    llm_temp_label.font = _BOLD
    llm_temp_label.alignment = top_aligned
    ws.cell(row=20, column=2, value=llm_meta.get("temperature"))

    llm_sha_label = ws.cell(row=21, column=1, value="llm_prompt_sha256")
    llm_sha_label.font = _BOLD
    llm_sha_label.alignment = top_aligned
    ws.cell(row=21, column=2, value=str(llm_meta.get("prompt_sha256", "")))

    llm_raw_label = ws.cell(row=22, column=1, value="llm_raw_response")
    llm_raw_label.font = _BOLD
    llm_raw_label.alignment = top_aligned
    llm_raw_value = ws.cell(row=22, column=2, value=verdict.raw_response or "")
    llm_raw_value.alignment = wrap_top_left

    # Collapse rows 17-22; +/- button next to row 16 (the visible header).
    ws.sheet_properties.outlinePr.summaryBelow = False
    for row_idx in range(17, 23):
        ws.row_dimensions[row_idx].outline_level = 1
        ws.row_dimensions[row_idx].hidden = True

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 110
    ws.freeze_panes = "B1"


def _load_meta(meta_path: Path | None) -> tuple[str, dict[str, object]]:
    """Return ``(criterion, llm_meta_dict)`` from a screening-meta.json file.

    Falls back to placeholders when ``meta_path`` is None or missing —
    the export still works, just without the criterion / LLM
    fingerprint. ``screening-meta.json`` may pre-date the addition of
    ``prompt`` to ``ScreeningMeta``, in which case the criterion shows
    a stub.
    """
    if meta_path is None or not meta_path.exists():
        return ("<screening criterion not available>", {})

    raw = json.loads(meta_path.read_text())
    # Use ScreeningMeta to validate, but tolerate older meta files.
    try:
        meta = ScreeningMeta.model_validate(raw)
        criterion = meta.prompt or "<screening criterion not recorded in meta>"
        llm_meta = {
            "model": meta.llm.model,
            "temperature": meta.llm.temperature,
            "prompt_sha256": meta.llm.prompt_sha256,
        }
        return (criterion, llm_meta)
    except Exception:
        criterion = raw.get("prompt") or "<screening criterion not recorded in meta>"
        llm_meta = raw.get("llm", {}) or {}
        return (criterion, llm_meta)


def export_review_xlsx(
    verdicts_path: Path,
    catalogue_path: Path,
    output_path: Path,
    *,
    n_subset: int | None,
    seed: int = _DEFAULT_SUBSET_SEED,
    meta_path: Path | None = None,
) -> int:
    """Build and save the review workbook, returning the per-work sheet count."""
    catalogue: dict[str, Work] = {w.id: w for w in read_jsonl(catalogue_path, Work)}
    verdicts = list(read_jsonl(verdicts_path, ScreeningVerdict))
    for verdict in verdicts:
        if verdict.work_id not in catalogue:
            raise ValueError(
                f"work_id {verdict.work_id!r} in {verdicts_path} not found in "
                f"catalogue {catalogue_path}"
            )

    selected = sample_verdicts(verdicts, n_subset, seed)
    criterion, llm_meta = _load_meta(meta_path)

    wb = Workbook()
    index_ws = wb.active
    if index_ws is None:
        raise RuntimeError("openpyxl Workbook() did not produce a default worksheet")

    used_names: set[str] = set()
    items: list[tuple[ScreeningVerdict, Work, str]] = []
    for verdict in selected:
        sheet_name = _unique_sheet_name(short_work_id(verdict.work_id), used_names)
        ws = wb.create_sheet(title=sheet_name)
        build_work_sheet(
            ws,
            verdict,
            catalogue[verdict.work_id],
            criterion=criterion,
            llm_meta=llm_meta,
        )
        items.append((verdict, catalogue[verdict.work_id], sheet_name))

    build_index_sheet(index_ws, items)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return len(selected)


# ── CLI ───────────────────────────────────────────────────────────────────────


def build_subparser(
    subparsers: argparse._SubParsersAction[argparse.ArgumentParser],
) -> argparse.ArgumentParser:
    parser = subparsers.add_parser(
        "screening-abstracts-export",
        help="Export stage 3 verdicts + catalogue to CSV or XLSX.",
    )
    parser.add_argument(
        "--format",
        choices=["csv", "xlsx"],
        required=True,
        help="Output format: 'csv' for flat review spreadsheet, 'xlsx' for per-work workbook.",
    )
    parser.add_argument(
        "--verdicts",
        type=Path,
        required=True,
        help="Path to verdicts.jsonl from screening-abstracts",
    )
    parser.add_argument(
        "--catalogue",
        type=Path,
        required=True,
        help="Path to the deduplicated catalogue JSONL",
    )
    parser.add_argument(
        "--meta",
        type=Path,
        default=None,
        help=(
            "Path to screening-meta.json. Default: <verdicts parent>/screening-meta.json. "
            "Used by --format xlsx to embed the screening criterion and LLM fingerprint."
        ),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help=(
            "Output file path. Default: <verdicts parent>/review.csv (CSV) or "
            "<verdicts parent>/review.xlsx (XLSX)."
        ),
    )
    parser.add_argument(
        "--n-subset",
        type=int,
        default=None,
        help="Random sample size; emit all when unset or >= total. XLSX only.",
    )
    parser.add_argument(
        "--subset-seed",
        type=int,
        default=_DEFAULT_SUBSET_SEED,
        help=f"Random seed for --n-subset (default: {_DEFAULT_SUBSET_SEED}). XLSX only.",
    )
    parser.set_defaults(run=run)
    return parser


def run(args: argparse.Namespace) -> None:
    verdicts_path: Path = args.verdicts

    if args.format == "csv":
        if args.n_subset is not None or args.subset_seed != _DEFAULT_SUBSET_SEED:
            sys.exit(
                "--n-subset and --subset-seed are only valid with --format xlsx"
            )
        output_path: Path = (
            args.output
            if args.output is not None
            else verdicts_path.parent / "review.csv"
        )
        count = export_review_csv(verdicts_path, args.catalogue, output_path)
        print(f"Wrote {count} rows to {output_path}", file=sys.stderr)

    else:  # xlsx
        output_path = (
            args.output
            if args.output is not None
            else verdicts_path.parent / "review.xlsx"
        )
        meta_path: Path = (
            args.meta
            if args.meta is not None
            else verdicts_path.parent / "screening-meta.json"
        )
        count = export_review_xlsx(
            verdicts_path,
            args.catalogue,
            output_path,
            n_subset=args.n_subset,
            seed=args.subset_seed,
            meta_path=meta_path,
        )
        print(
            f"Wrote workbook with {count} per-work sheets to {output_path}",
            file=sys.stderr,
        )
