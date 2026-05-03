"""Tests for screening_abstracts.export (CSV and XLSX formats)."""

from __future__ import annotations

import csv
from pathlib import Path

import pytest
from openpyxl import load_workbook
from pydantic import BaseModel

from laglitsynth.screening_abstracts.export import (
    COLUMNS,
    _unique_sheet_name,
    build_row,
    export_review_csv,
    export_review_xlsx,
    sample_verdicts,
    short_work_id,
)
from laglitsynth.screening_abstracts.models import ScreeningVerdict

from conftest import _make_work


def _write_jsonl(path: Path, records: list[BaseModel]) -> None:
    with open(path, "w") as f:
        for r in records:
            f.write(r.model_dump_json() + "\n")


def _write_inputs(
    tmp_path: Path,
    works: list[Work],
    verdicts: list[ScreeningVerdict],
    *,
    output_suffix: str = "csv",
) -> tuple[Path, Path, Path]:
    catalogue_path = tmp_path / "dedup.jsonl"
    verdicts_path = tmp_path / "verdicts.jsonl"
    output_path = tmp_path / f"review.{output_suffix}"
    _write_jsonl(catalogue_path, list(works))
    _write_jsonl(verdicts_path, list(verdicts))
    return verdicts_path, catalogue_path, output_path


# ── CSV ───────────────────────────────────────────────────────────────────────


def test_export_round_trip(tmp_path: Path) -> None:
    works = [_make_work("W1"), _make_work("W2", abstract=None)]
    verdicts = [
        ScreeningVerdict(
            work_id="W1",
            relevance_score=80,
            reason="relevant",
            seed=42,
            raw_response='{"relevance_score": 80, "reason": "relevant"}',
        ),
        ScreeningVerdict(
            work_id="W2",
            relevance_score=None,
            reason="no-abstract",
            seed=None,
            raw_response=None,
        ),
    ]
    verdicts_path, catalogue_path, output_path = _write_inputs(tmp_path, works, verdicts)

    count = export_review_csv(verdicts_path, catalogue_path, output_path)

    assert count == 2
    with open(output_path, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    assert [r["work_id"] for r in rows] == ["W1", "W2"]
    assert rows[0]["relevance_score"] == "80"
    assert rows[0]["llm_reason"] == "relevant"
    assert rows[0]["reviewer_decision"] == ""
    assert rows[0]["reviewer_reason"] == ""
    assert rows[0]["abstract"] == "An abstract."
    # sentinel — empty relevance_score cell, sentinel reason preserved
    assert rows[1]["relevance_score"] == ""
    assert rows[1]["llm_reason"] == "no-abstract"
    assert rows[1]["abstract"] == ""


def test_export_special_characters_survive(tmp_path: Path) -> None:
    tricky_abstract = 'Line 1, with comma\nLine 2 with "quotes" and, more commas'
    works = [_make_work("W1", abstract=tricky_abstract)]
    verdicts = [
        ScreeningVerdict(
            work_id="W1",
            relevance_score=50,
            reason='Uses "X, Y"',
            seed=1,
        )
    ]
    verdicts_path, catalogue_path, output_path = _write_inputs(tmp_path, works, verdicts)

    export_review_csv(verdicts_path, catalogue_path, output_path)

    with open(output_path, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["abstract"] == tricky_abstract
    assert rows[0]["llm_reason"] == 'Uses "X, Y"'


def test_export_raises_on_missing_work(tmp_path: Path) -> None:
    works = [_make_work("W1")]
    verdicts = [
        ScreeningVerdict(work_id="W_missing", relevance_score=None, reason="?"),
    ]
    verdicts_path, catalogue_path, output_path = _write_inputs(tmp_path, works, verdicts)

    with pytest.raises(ValueError, match="W_missing"):
        export_review_csv(verdicts_path, catalogue_path, output_path)


def test_export_column_order(tmp_path: Path) -> None:
    works = [_make_work("W1")]
    verdicts = [ScreeningVerdict(work_id="W1", relevance_score=80, reason="ok", seed=1)]
    verdicts_path, catalogue_path, output_path = _write_inputs(tmp_path, works, verdicts)

    export_review_csv(verdicts_path, catalogue_path, output_path)

    with open(output_path, encoding="utf-8-sig", newline="") as f:
        header = next(csv.reader(f))
    assert tuple(header) == COLUMNS


def test_export_writes_utf8_bom(tmp_path: Path) -> None:
    works = [_make_work("W1")]
    verdicts = [ScreeningVerdict(work_id="W1", relevance_score=80, reason="ok", seed=1)]
    verdicts_path, catalogue_path, output_path = _write_inputs(tmp_path, works, verdicts)

    export_review_csv(verdicts_path, catalogue_path, output_path)

    with open(output_path, "rb") as f:
        head = f.read(3)
    assert head == b"\xef\xbb\xbf"


def test_build_row_empty_fields() -> None:
    work = _make_work(
        "W1", title=None, doi=None, publication_year=None, abstract=None
    )
    verdict = ScreeningVerdict(work_id="W1", relevance_score=None, reason=None)
    row = build_row(verdict, work)
    assert row["title"] == ""
    assert row["doi"] == ""
    assert row["publication_year"] == ""
    assert row["abstract"] == ""
    assert row["relevance_score"] == ""
    assert row["llm_reason"] == ""
    assert row["raw_response"] == ""


# ── XLSX ──────────────────────────────────────────────────────────────────────


def test_sample_verdicts_none_returns_all() -> None:
    vs = [ScreeningVerdict(work_id=f"W{i}") for i in range(5)]
    assert sample_verdicts(vs, None, seed=0) == vs


def test_sample_verdicts_oversize_returns_all() -> None:
    vs = [ScreeningVerdict(work_id=f"W{i}") for i in range(3)]
    assert sample_verdicts(vs, 10, seed=0) == vs
    assert sample_verdicts(vs, 3, seed=0) == vs


def test_sample_verdicts_preserves_original_order() -> None:
    vs = [ScreeningVerdict(work_id=f"W{i}") for i in range(10)]
    sample = sample_verdicts(vs, 5, seed=42)
    assert len(sample) == 5
    # Emitted verdicts appear in the same relative order as in vs.
    positions = [vs.index(v) for v in sample]
    assert positions == sorted(positions)


def test_sample_verdicts_reproducible() -> None:
    vs = [ScreeningVerdict(work_id=f"W{i}") for i in range(20)]
    a = [v.work_id for v in sample_verdicts(vs, 7, seed=123)]
    b = [v.work_id for v in sample_verdicts(vs, 7, seed=123)]
    c = [v.work_id for v in sample_verdicts(vs, 7, seed=456)]
    assert a == b
    assert a != c


def test_short_work_id_openalex_url() -> None:
    assert short_work_id("https://openalex.org/W3213722062") == "W3213722062"


def test_short_work_id_truncates_long() -> None:
    long_id = "A" * 100
    assert short_work_id(long_id) == "A" * 31


def test_short_work_id_strips_forbidden_chars() -> None:
    # Excel forbids / \ ? * [ ] : in sheet names.
    assert short_work_id("prefix/W[foo]:bar") == "W_foo__bar"


def test_unique_sheet_name_collision_suffix() -> None:
    used: set[str] = set()
    assert _unique_sheet_name("W1", used) == "W1"
    assert _unique_sheet_name("W1", used) == "W1_2"
    assert _unique_sheet_name("W1", used) == "W1_3"


def test_unique_sheet_name_truncates_base_to_fit_suffix() -> None:
    used: set[str] = {"A" * 31}
    result = _unique_sheet_name("A" * 31, used)
    # Base gets trimmed to make room for "_2" suffix; total still <= 31.
    assert len(result) == 31
    assert result.endswith("_2")


def test_xlsx_export_filters_no_abstract(tmp_path: Path) -> None:
    """no-abstract verdicts are dropped — unscoreable for the reviewer.

    llm-parse-failure and llm-timeout still appear (they have an
    abstract; only the LLM call failed)."""
    works = [
        _make_work("https://openalex.org/W1"),
        _make_work("https://openalex.org/W2", abstract=None),
        _make_work("https://openalex.org/W3"),
        _make_work("https://openalex.org/W4"),
    ]
    verdicts = [
        ScreeningVerdict(
            work_id="https://openalex.org/W1",
            relevance_score=80,
            reason="relevant",
            seed=1,
        ),
        ScreeningVerdict(
            work_id="https://openalex.org/W2",
            relevance_score=None,
            reason="no-abstract",
            seed=None,
        ),
        ScreeningVerdict(
            work_id="https://openalex.org/W3",
            relevance_score=None,
            reason="llm-parse-failure",
            seed=None,
            raw_response="malformed",
        ),
        ScreeningVerdict(
            work_id="https://openalex.org/W4",
            relevance_score=None,
            reason="llm-timeout",
            seed=None,
        ),
    ]
    verdicts_path, catalogue_path, output_path = _write_inputs(
        tmp_path, works, verdicts, output_suffix="xlsx"
    )

    count = export_review_xlsx(
        verdicts_path, catalogue_path, output_path, n_subset=None
    )

    # W2 (no-abstract) is dropped; W1, W3, W4 stay.
    assert count == 3
    wb = load_workbook(output_path)
    assert wb.sheetnames == ["Index", "W1", "W3", "W4"]
    index = wb["Index"]
    # Reviewer-identity rows above the table.
    assert index["A1"].value == "reviewer_name"
    assert index["A2"].value == "reviewer_email"
    assert index["A3"].value == "review_date"
    # Header row at row 5; data rows from row 6.
    assert index["A5"].value == "work_id"
    assert index["B5"].value == "title"
    assert index["C5"].value == "authors"
    assert index["D5"].value == "journal"
    assert index["E5"].value == "year"
    assert index["F5"].value == "sheet"
    # No LLM columns on the index — the reviewer scores blind.
    assert index["G5"].value is None
    assert index["A6"].value == "https://openalex.org/W1"
    assert index["A7"].value == "https://openalex.org/W3"
    assert index["A8"].value == "https://openalex.org/W4"
    # Index hyperlink points at the per-work sheet via the canonical
    # in-workbook ``location`` form.
    link = index["F6"]
    assert link.value == "W1"
    assert link.hyperlink is not None
    assert link.hyperlink.location == "'W1'!A1"


def test_xlsx_export_subset_preserves_verdict_order(tmp_path: Path) -> None:
    works = [_make_work(f"https://openalex.org/W{i}") for i in range(10)]
    verdicts = [
        ScreeningVerdict(
            work_id=f"https://openalex.org/W{i}",
            relevance_score=i * 10,
            reason=f"reason-{i}",
            seed=i,
        )
        for i in range(10)
    ]
    verdicts_path, catalogue_path, output_path = _write_inputs(
        tmp_path, works, verdicts, output_suffix="xlsx"
    )

    count = export_review_xlsx(
        verdicts_path, catalogue_path, output_path, n_subset=4, seed=7
    )

    assert count == 4
    wb = load_workbook(output_path)
    assert wb.sheetnames[0] == "Index"
    per_work = wb.sheetnames[1:]
    assert len(per_work) == 4
    # Sheet names (W0..W9) must appear in ascending numeric order.
    indices = [int(name[1:]) for name in per_work]
    assert indices == sorted(indices)


def test_xlsx_export_subset_oversize_emits_full_set(tmp_path: Path) -> None:
    works = [_make_work(f"https://openalex.org/W{i}") for i in range(3)]
    verdicts = [
        ScreeningVerdict(
            work_id=f"https://openalex.org/W{i}",
            relevance_score=50,
            reason="ok",
            seed=i,
        )
        for i in range(3)
    ]
    verdicts_path, catalogue_path, output_path = _write_inputs(
        tmp_path, works, verdicts, output_suffix="xlsx"
    )

    count = export_review_xlsx(
        verdicts_path, catalogue_path, output_path, n_subset=100
    )

    assert count == 3
    wb = load_workbook(output_path)
    assert wb.sheetnames == ["Index", "W0", "W1", "W2"]


def test_xlsx_export_llm_failure_sentinels_keep_sheet(tmp_path: Path) -> None:
    """llm-parse-failure / llm-timeout still get a per-work sheet.

    The abstract is present so the reviewer can score; only the LLM
    output is unusable. ``llm_score`` is blank, ``llm_reason`` carries
    the sentinel string.
    """
    works = [_make_work("https://openalex.org/W1")]
    verdicts = [
        ScreeningVerdict(
            work_id="https://openalex.org/W1",
            relevance_score=None,
            reason="llm-timeout",
            seed=None,
        )
    ]
    verdicts_path, catalogue_path, output_path = _write_inputs(
        tmp_path, works, verdicts, output_suffix="xlsx"
    )

    export_review_xlsx(verdicts_path, catalogue_path, output_path, n_subset=None)

    wb = load_workbook(output_path)
    work_sheet = wb["W1"]
    # llm_score is blank for sentinel verdicts.
    assert work_sheet["A18"].value == "llm_score"
    assert work_sheet["B18"].value is None
    # llm_reason carries the sentinel string.
    assert work_sheet["A19"].value == "llm_reason"
    assert work_sheet["B19"].value == "llm-timeout"


def test_xlsx_work_sheet_layout(tmp_path: Path) -> None:
    """Per-work sheet has the new layout: biblio block, criterion, scoring,
    collapsed LLM details with hyperlinks and percent format."""
    work = _make_work(
        "https://openalex.org/W1",
        title="The Title",
        abstract="An abstract.",
        doi="10.1000/xyz",
        publication_year=2024,
    )
    verdicts = [
        ScreeningVerdict(
            work_id="https://openalex.org/W1",
            relevance_score=85,
            reason="relevant",
            seed=1,
            raw_response='{"relevance_score": 85, "reason": "relevant"}',
        )
    ]
    verdicts_path, catalogue_path, output_path = _write_inputs(
        tmp_path, [work], verdicts, output_suffix="xlsx"
    )

    # Write a meta file so the criterion is embedded.
    meta_path = tmp_path / "screening-meta.json"
    meta_path.write_text(
        '{"run": {"tool": "t", "tool_version": "alpha", "run_at": "2026-01-01T00:00:00.000000+00:00", "validation_skipped": 0},'
        ' "llm": {"model": "gemma3:4b", "temperature": 0.8, "prompt_sha256": "' + ("a" * 64) + '"},'
        ' "threshold": 50, "input_path": "x", "input_count": 1,'
        ' "above_threshold_count": 1, "below_threshold_count": 0, "skipped_count": 0,'
        ' "llm_parse_failure_count": 0, "llm_timeout_count": 0,'
        ' "prompt": "On a scale from 0% to 100%, how relevant?"}'
    )

    export_review_xlsx(
        verdicts_path,
        catalogue_path,
        output_path,
        n_subset=None,
        meta_path=meta_path,
    )
    wb = load_workbook(output_path)
    ws = wb["W1"]

    # Row 1: back-to-Index hyperlink.
    assert ws["A1"].value == "← back to Index"
    assert ws["A1"].hyperlink is not None
    assert ws["A1"].hyperlink.location == "'Index'!A1"

    # Bibliographic block (rows 3-9).
    assert ws["A3"].value == "title"
    assert ws["B3"].value == "The Title"
    assert ws["A4"].value == "authors"
    assert ws["A5"].value == "journal"
    assert ws["A6"].value == "publication_year"
    assert ws["B6"].value == 2024
    assert ws["A7"].value == "doi"
    assert ws["B7"].value == "10.1000/xyz"
    # DOI is an external hyperlink.
    assert ws["B7"].hyperlink is not None
    assert ws["B7"].hyperlink.target == "https://doi.org/10.1000/xyz"
    assert ws["A8"].value == "openalex"
    # OpenAlex URL is an external hyperlink.
    assert ws["B8"].hyperlink is not None
    assert ws["B8"].hyperlink.target == "https://openalex.org/W1"
    assert ws["A9"].value == "abstract"
    assert ws["B9"].value == "An abstract."

    # Criterion + scoring + remarks block.
    assert ws["A11"].value == "criterion"
    assert ws["B11"].value == "On a scale from 0% to 100%, how relevant?"
    assert ws["A12"].value == "scoring_instructions"
    assert ws["A13"].value == "reviewer_score"
    assert "<insert relevance" in ws["B13"].value
    assert ws["A14"].value == "reviewer_reason"
    assert ws["A15"].value == "reviewer_remarks"
    assert "<any general feedback" in ws["B15"].value

    # LLM details header (row 17, visible).
    assert ws["A17"].value == "LLM details (expand to peek)"
    # Collapsed LLM rows with outline level 1.
    for row_idx in range(18, 24):
        assert ws.row_dimensions[row_idx].outline_level == 1
        assert ws.row_dimensions[row_idx].hidden is True
    # summaryBelow=False: button at the visible header row.
    assert ws.sheet_properties.outlinePr.summaryBelow is False

    # llm_score formatted as percent (0.85).
    assert ws["A18"].value == "llm_score"
    assert ws["B18"].value == 0.85
    assert ws["B18"].number_format == "0%"
    assert ws["A19"].value == "llm_reason"
    assert ws["B19"].value == "relevant"
    assert ws["A20"].value == "llm_model"
    assert ws["B20"].value == "gemma3:4b"
    assert ws["A21"].value == "llm_temperature"
    assert ws["A22"].value == "llm_prompt_sha256"
    assert ws["A23"].value == "llm_raw_response"


def test_xlsx_export_raises_on_missing_work(tmp_path: Path) -> None:
    works = [_make_work("https://openalex.org/W1")]
    verdicts = [
        ScreeningVerdict(
            work_id="https://openalex.org/W_missing",
            relevance_score=None,
            reason="?",
        )
    ]
    verdicts_path, catalogue_path, output_path = _write_inputs(
        tmp_path, works, verdicts, output_suffix="xlsx"
    )

    with pytest.raises(ValueError, match="W_missing"):
        export_review_xlsx(
            verdicts_path, catalogue_path, output_path, n_subset=None
        )


# ── CLI smoke ─────────────────────────────────────────────────────────────────


def test_cli_csv_default_output(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    from laglitsynth.cli import main

    works = [_make_work("W1")]
    verdicts = [ScreeningVerdict(work_id="W1", relevance_score=80, reason="ok", seed=1)]
    verdicts_path, catalogue_path, _ = _write_inputs(tmp_path, works, verdicts)

    main(
        [
            "screening-abstracts-export",
            "--format",
            "csv",
            "--verdicts",
            str(verdicts_path),
            "--catalogue",
            str(catalogue_path),
        ]
    )

    # Default resolves to <verdicts parent>/review.csv
    assert (tmp_path / "review.csv").exists()


def test_cli_csv_custom_output(tmp_path: Path) -> None:
    from laglitsynth.cli import main

    works = [_make_work("W1")]
    verdicts = [ScreeningVerdict(work_id="W1", relevance_score=80, reason="ok", seed=1)]
    verdicts_path, catalogue_path, _ = _write_inputs(tmp_path, works, verdicts)
    custom = tmp_path / "custom" / "my-review.csv"

    main(
        [
            "screening-abstracts-export",
            "--format",
            "csv",
            "--verdicts",
            str(verdicts_path),
            "--catalogue",
            str(catalogue_path),
            "--output",
            str(custom),
        ]
    )

    assert custom.exists()


def test_cli_csv_rejects_n_subset(tmp_path: Path) -> None:
    """--n-subset with --format csv must fail with a clear error."""
    from laglitsynth.cli import main

    works = [_make_work("W1")]
    verdicts = [ScreeningVerdict(work_id="W1", relevance_score=80, reason="ok", seed=1)]
    verdicts_path, catalogue_path, _ = _write_inputs(tmp_path, works, verdicts)

    with pytest.raises(SystemExit) as exc_info:
        main(
            [
                "screening-abstracts-export",
                "--format",
                "csv",
                "--verdicts",
                str(verdicts_path),
                "--catalogue",
                str(catalogue_path),
                "--n-subset",
                "10",
            ]
        )

    assert exc_info.value.code == "--n-subset and --subset-seed are only valid with --format xlsx"


def test_cli_xlsx_default_output(tmp_path: Path) -> None:
    from laglitsynth.cli import main

    works = [_make_work("https://openalex.org/W1")]
    verdicts = [
        ScreeningVerdict(
            work_id="https://openalex.org/W1",
            relevance_score=80,
            reason="ok",
            seed=1,
        )
    ]
    verdicts_path, catalogue_path, _ = _write_inputs(
        tmp_path, works, verdicts, output_suffix="xlsx"
    )

    main(
        [
            "screening-abstracts-export",
            "--format",
            "xlsx",
            "--verdicts",
            str(verdicts_path),
            "--catalogue",
            str(catalogue_path),
        ]
    )

    assert (tmp_path / "review.xlsx").exists()


def test_cli_xlsx_subset(tmp_path: Path) -> None:
    """--format xlsx --n-subset N produces a per-work-tab workbook with N sheets."""
    from laglitsynth.cli import main

    works = [_make_work(f"https://openalex.org/W{i}") for i in range(6)]
    verdicts = [
        ScreeningVerdict(
            work_id=f"https://openalex.org/W{i}",
            relevance_score=60,
            reason="ok",
            seed=i,
        )
        for i in range(6)
    ]
    verdicts_path, catalogue_path, _ = _write_inputs(
        tmp_path, works, verdicts, output_suffix="xlsx"
    )
    output = tmp_path / "sub.xlsx"

    main(
        [
            "screening-abstracts-export",
            "--format",
            "xlsx",
            "--verdicts",
            str(verdicts_path),
            "--catalogue",
            str(catalogue_path),
            "--output",
            str(output),
            "--n-subset",
            "2",
            "--subset-seed",
            "3",
        ]
    )

    wb = load_workbook(output)
    # Index + 2 per-work sheets.
    assert len(wb.sheetnames) == 3
